"""手数料計算 API ルーター。

prefix: /api/v1/commission

Endpoints:
- POST /upload          ファイル複数アップロード → セッション ID 発行
- POST /calculate/{sid} SSE で計算進捗配信、interrupt で HITL に停止
- GET  /results/{sid}   計算結果取得 (フィルタ・ページング)
- POST /hitl/approve/{sid} HITL 承認 → グラフ再開
- GET  /export/{sid}    確定済みデータを xlsx で出力
- GET  /dashboard/{sid} KPI / 取引先別・商材別サマリ
"""

from __future__ import annotations

import io
import logging
from typing import AsyncIterator

from datarobot.auth.session import AuthCtx
from datarobot.auth.typing import Metadata
from fastapi import APIRouter, Depends, File, HTTPException, Request, UploadFile, status
from fastapi.responses import StreamingResponse

from app.auth.ctx import must_get_auth_ctx
from app.commission.graph_driver import (
    resume_with_decisions,
    run_calculate,
    stream_to_sse,
)
from app.commission.schemas import (
    CalculateOptions,
    DashboardKPI,
    DashboardResponse,
    HitlApproveRequest,
    HitlApproveResponse,
    ResultsResponse,
    UploadedFileInfo,
    UploadResponse,
)
from app.commission.session_store import CommissionSessionStore
from app.deps import Deps

logger = logging.getLogger(__name__)
commission_router = APIRouter(prefix="/commission", tags=["Commission"])


def _detect_type(filename: str) -> str:
    if "売上明細" in filename or "売上" in filename:
        return "sales"
    if "取引条件" in filename or "マスタ" in filename:
        return "master"
    return "unknown"


def _get_store(request: Request) -> CommissionSessionStore:
    deps: Deps = request.app.state.deps
    store = getattr(deps, "commission_session_store", None)
    if store is None:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="commission session store not initialized",
        )
    return store


@commission_router.post("/upload", response_model=UploadResponse)
async def upload_files(
    request: Request,
    files: list[UploadFile] = File(...),
    _auth: AuthCtx[Metadata] = Depends(must_get_auth_ctx),
) -> UploadResponse:
    """Excel ファイル複数を受け取って一時保存し、セッション ID を返す。"""
    store = _get_store(request)
    sess = await store.create()

    uploaded: list[UploadedFileInfo] = []
    for f in files:
        content = await f.read()
        filename = f.filename or "unknown.xlsx"
        detected = _detect_type(filename)
        rec = await store.add_file(
            session_id=sess.session_id,
            filename=filename,
            content=content,
            detected_type=detected,
        )
        uploaded.append(
            UploadedFileInfo(
                file_id=rec.file_id,
                filename=rec.filename,
                size=rec.size,
                detected_type=detected,  # type: ignore[arg-type]
            )
        )

    return UploadResponse(
        session_id=sess.session_id,
        uploaded=uploaded,
        message=f"{len(uploaded)} ファイル受領、計算可能です",
    )


@commission_router.post("/calculate/{session_id}")
async def calculate(
    request: Request,
    session_id: str,
    options: CalculateOptions | None = None,
    _auth: AuthCtx[Metadata] = Depends(must_get_auth_ctx),
) -> StreamingResponse:
    """SSE でグラフ実行進捗を配信。"""
    store = _get_store(request)
    sess = await store.get(session_id)
    if sess is None:
        raise HTTPException(status_code=404, detail="session not found")

    threshold = (options or CalculateOptions()).anomaly_threshold

    files = [
        {
            "file_id": f.file_id,
            "filename": f.filename,
            "path": f.path,
            "detected_type": f.detected_type,
        }
        for f in sess.uploaded_files
    ]

    async def event_stream() -> AsyncIterator[str]:
        # SSE 配信
        async for sse_str in stream_to_sse(
            run_calculate(
                session_id=session_id,
                uploaded_files=files,
                threshold=threshold,
            )
        ):
            yield sse_str
        # ストリーム終了後にセッションのスナップショットを更新
        from app.commission_engine.graph import get_checkpointer  # noqa: WPS433

        try:
            cp = get_checkpointer()
            snap = cp.get(  # type: ignore[attr-defined]
                {"configurable": {"thread_id": session_id}}
            )
            if snap and "channel_values" in snap:
                values = snap["channel_values"]
                await store.update_results(
                    session_id,
                    calculation_results=values.get("calculation_results") or [],
                    pending_hitl=values.get("pending_hitl") or [],
                    summary=values.get("summary") or {},
                    processing_status=values.get("processing_status", "review"),
                )
        except Exception as e:
            logger.warning("post-run state sync failed: %s", e)

    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={"X-Accel-Buffering": "no", "Cache-Control": "no-cache"},
    )


@commission_router.get("/results/{session_id}", response_model=ResultsResponse)
async def get_results(
    request: Request,
    session_id: str,
    status_filter: str = "all",
    page: int = 1,
    per_page: int = 50,
    _auth: AuthCtx[Metadata] = Depends(must_get_auth_ctx),
) -> ResultsResponse:
    """計算結果をフィルタ・ページングで取得。

    status_filter: all / hitl_pending / approved / error
    """
    store = _get_store(request)
    sess = await store.get(session_id)
    if sess is None:
        raise HTTPException(status_code=404, detail="session not found")

    if status_filter == "hitl_pending":
        rows = sess.pending_hitl
    elif status_filter == "approved":
        rows = sess.approved_results
    elif status_filter == "error":
        rows = [r for r in sess.calculation_results if r.get("error_message")]
    else:
        rows = sess.calculation_results

    total = len(rows)
    start = max(0, (page - 1) * per_page)
    end = start + per_page
    paged = rows[start:end]

    return ResultsResponse(
        results=paged,
        total=total,
        page=page,
        per_page=per_page,
        summary=sess.summary or None,
    )


@commission_router.post(
    "/hitl/approve/{session_id}", response_model=HitlApproveResponse
)
async def approve_hitl(
    request: Request,
    session_id: str,
    body: HitlApproveRequest,
    _auth: AuthCtx[Metadata] = Depends(must_get_auth_ctx),
) -> HitlApproveResponse:
    """HITL 承認結果でグラフを再開して approved_results を確定。"""
    store = _get_store(request)
    sess = await store.get(session_id)
    if sess is None:
        raise HTTPException(status_code=404, detail="session not found")

    decisions = [d.model_dump() for d in body.approvals]

    # グラフを再開（ストリームは消費するだけ）
    async for _ev, _data in resume_with_decisions(
        session_id=session_id, decisions=decisions
    ):
        pass

    # 再開後のスナップショットを反映
    try:
        from app.commission_engine.graph import get_checkpointer  # noqa: WPS433

        cp = get_checkpointer()
        snap = cp.get(  # type: ignore[attr-defined]
            {"configurable": {"thread_id": session_id}}
        )
        if snap and "channel_values" in snap:
            values = snap["channel_values"]
            await store.update_results(
                session_id,
                approved_results=values.get("approved_results") or [],
                processing_status=values.get("processing_status", "approved"),
            )
    except Exception as e:
        logger.warning("approve snapshot sync failed: %s", e)

    approved_count = sum(1 for d in body.approvals if d.action == "approve")
    rejected_count = sum(1 for d in body.approvals if d.action == "reject")
    manual_count = sum(1 for d in body.approvals if d.action == "manual")
    remaining = max(0, len(sess.pending_hitl) - len(body.approvals))

    return HitlApproveResponse(
        approved_count=approved_count,
        rejected_count=rejected_count,
        manual_count=manual_count,
        remaining_hitl=remaining,
    )


@commission_router.get("/export/{session_id}")
async def export_excel(
    request: Request,
    session_id: str,
    _auth: AuthCtx[Metadata] = Depends(must_get_auth_ctx),
) -> StreamingResponse:
    """確定済みデータを Excel (.xlsx) で出力。明細 + サマリの 2 シート構成。"""
    store = _get_store(request)
    sess = await store.get(session_id)
    if sess is None:
        raise HTTPException(status_code=404, detail="session not found")

    rows = sess.approved_results or sess.calculation_results
    if not rows:
        raise HTTPException(status_code=400, detail="no data to export")

    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    wb = openpyxl.Workbook()
    # 明細シート
    ws = wb.active
    ws.title = "明細"
    headers = [
        "レコードNo",
        "取引先コード",
        "取引先名称",
        "商材",
        "決済方法",
        "基本コミッション",
        "ボリュームインセン",
        "特別コミッション①",
        "特別コミッション②",
        "継続コミッション",
        "紹介制度コミッション",
        "PAPコミッション",
        "PASコミッション",
        "PHコミッション",
        "QI分割額",
        "口振初回手数料",
        "戻入手数料",
        "合計手数料",
        "ステータス",
        "計算根拠",
    ]
    ws.append(headers)
    for r in rows:
        ws.append(
            [
                r.get("record_no"),
                r.get("partner_code"),
                r.get("partner_name"),
                r.get("product"),
                r.get("payment_method"),
                r.get("basic_commission"),
                r.get("volume_incentive"),
                r.get("special_commission_1"),
                r.get("special_commission_2"),
                r.get("continuous_commission"),
                r.get("referral_commission"),
                r.get("pap_commission"),
                r.get("pas_commission"),
                r.get("ph_commission"),
                r.get("qi_amount"),
                r.get("debit_initial_fee"),
                r.get("return_amount"),
                r.get("total_commission"),
                "確定" if not r.get("is_anomaly") else r.get("hitl_reason"),
                "\n".join(r.get("calculation_trace", [])),
            ]
        )
    # 列幅自動調整
    for col_idx, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(
            12, min(40, len(h) * 2)
        )

    # サマリシート
    ws2 = wb.create_sheet("サマリ")
    summary = sess.summary or {}
    ws2.append(["項目", "値"])
    ws2.append(["総レコード数", summary.get("total_records", 0)])
    ws2.append(["自動完了", summary.get("auto_completed", 0)])
    ws2.append(["HITL対象", summary.get("hitl_pending", 0)])
    ws2.append(["合計手数料 (円)", summary.get("total_commission_amount", 0)])
    ws2.append(
        ["自動完了率", f"{summary.get('auto_completion_rate', 0) * 100:.2f}%"]
    )
    ws2.append([])
    ws2.append(["取引先別合計"])
    ws2.append(["取引先名称", "件数", "合計手数料"])
    for p in summary.get("by_partner", []):
        ws2.append([p.get("name"), p.get("count"), p.get("total")])
    ws2.append([])
    ws2.append(["商材別合計"])
    ws2.append(["商材名", "件数", "合計手数料"])
    for p in summary.get("by_product", []):
        ws2.append([p.get("name"), p.get("count"), p.get("total")])
    for col_idx in range(1, 4):
        ws2.column_dimensions[get_column_letter(col_idx)].width = 24

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"commission_{session_id[:8]}.xlsx"
    return StreamingResponse(
        buf,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@commission_router.get("/dashboard/{session_id}", response_model=DashboardResponse)
async def get_dashboard(
    request: Request,
    session_id: str,
    _auth: AuthCtx[Metadata] = Depends(must_get_auth_ctx),
) -> DashboardResponse:
    """KPI + 取引先別 / 商材別サマリ。"""
    store = _get_store(request)
    sess = await store.get(session_id)
    if sess is None:
        raise HTTPException(status_code=404, detail="session not found")

    summary = sess.summary or {}
    kpi = DashboardKPI(
        total_records=summary.get("total_records", 0),
        auto_completed=summary.get("auto_completed", 0),
        hitl_pending=len(sess.pending_hitl),
        hitl_approved=len(sess.approved_results),
        error_count=summary.get("error_count", 0),
        total_commission_amount=summary.get("total_commission_amount", 0),
        auto_completion_rate=summary.get("auto_completion_rate", 0.0),
    )
    return DashboardResponse(
        kpi=kpi,
        by_partner=summary.get("by_partner", []),
        by_product=summary.get("by_product", []),
        processing_status=sess.processing_status,
    )
