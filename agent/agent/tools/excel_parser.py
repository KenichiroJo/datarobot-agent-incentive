# Copyright 2026 DataRobot, Inc.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#   http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
"""売上明細 Excel と取引条件マスタ Excel のパーサー。

サンプルデータ検証 (2026-05) で確認した実態:
- 売上明細: 192 列。商材列は粒度が粗いため「（Rename）商材」を使う
- 取引条件マスタ: 55 列。複合キーは「キー」列に事前計算済み
  (フォーマット: {一次店コード:10}{獲得月YYYYMMDD}{商材}{決済方法})
- 申込月 / 獲得月は openpyxl 経由で datetime.datetime として読み込まれる
"""
from __future__ import annotations

import logging
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

import openpyxl

logger = logging.getLogger(__name__)

# 売上明細から保持する列（spec 指定 + 計算に必要な列）
# 注: spec には「申込月、商材、決済方法、獲得者ID、（Rename）取引先コード」と並んでいるが、
#     サンプル検証の結果、突合キーには（Rename）系を使う必要があると判明したため両方保持する。
_SALES_COLUMNS_TO_KEEP: tuple[str, ...] = (
    "No",
    "ファイル区分",
    "レコードNo（手数料明細用）",
    "レコードNo",
    "契約ID",
    "申込日付",
    "申込月",
    "出荷日",
    "商材",
    "（Rename）商材",
    "出荷時決済方法",
    "（Rename）決済方法",
    "獲得者ID",
    "（Rename）取引先コード",
    "（Rename）取引先",
    "（Rename）取引区分",
    "（Rename）コミッション区分",
    "CSID",
    "獲得者名",
    "獲得店舗名",
    "配送個数",
    "販売価格_顧客",
    "基本コミッション",
    "ボリュームインセン",
    "特別コミッション",
    "特別コミッション2",
    "QI適用範囲",
    "分割計上期間",
    "口振初回手数料",
    "紹介制度区分",
    "紹介制度コミッション",
    "25ヶ月以降適用継続コミッション",
    "継続コミッション",
    "PAP区分",
    "PAPコミッション",
    "PAS区分",
    "PASコミッション",
    "PH区分",
    "PHコミッション",
    "違約金",
    "解約日",
    "請求ステータス",
    "営業販路",
)

# 取引条件マスタから保持する列
_MASTER_COLUMNS_TO_KEEP: tuple[str, ...] = (
    "キー",
    "取引先名称",
    "一次店コード",
    "一次店コード名称",
    "コミッション派生先コード",
    "コミッション派生先取引先名称",
    "コミッション派生先取引区分",
    "獲得月",
    "取引区分",
    "コミッション区分",
    "条件適用定義",
    "支払定義",
    "商材",
    "決済方法",
    "口振分割フラグ",
    "基本コミッション",
    "ボリュームインセンティブ",
    "配送周期インセンフラグ",
    "特別コミッション",
    "特別コミッション②",
    "QI適用範囲",
    "QI分割計上期間",
    "口振分割時初回手数料",
    "紹介制度区分",
    "紹介制度コミッション",
    "25・37ヶ月目以降継続コミッションフラグ",
    "継続コミッション",
    "PAP区分",
    "PAPコミッション",
    "PAS区分",
    "PASコミッション",
    "6L区分",  # スペックの PH区分 はマスタ側では 6L である可能性が高い
    "6Lコミッション",
    "デリキチ区分",
    "デリキチコミッション",
    "初回登録事務手数料",
    "保証金",
    "戻入条件",
    "戻入全額条件",
    "戻入半額条件",
    "違約金",
    "初回無料ボトル",
    "適用開始日",
    "適用終了日",
)


def normalize_yyyymmdd(value: Any) -> str | None:
    """申込月 / 獲得月の値を ``YYYYMMDD`` 形式の文字列に正規化する。

    対応する入力型:
    - ``datetime.datetime`` / ``datetime.date``: ``strftime("%Y%m%d")`` 相当
    - ``int`` / ``float``: Excel シリアル値（1900-01-01 起算）として変換
    - ``str``: ``YYYY-MM-DD`` / ``YYYY/MM/DD`` / ``YYYYMMDD`` / ``YYYYMM`` を吸収

    サンプル検証では 100% datetime だったが、本番運用での型ブレに耐性を持たせる。
    """
    if value is None:
        return None

    # datetime / date 系
    if isinstance(value, datetime):
        return value.strftime("%Y%m%d")
    if hasattr(value, "strftime"):
        # datetime.date など
        try:
            return value.strftime("%Y%m%d")
        except Exception:  # noqa: BLE001
            pass

    # Excel シリアル値（数値）→ 1900-01-01 起算の日数
    # Excel は 1900-02-29 を誤って閏日扱いするため -2 補正
    if isinstance(value, (int, float)):
        try:
            base = datetime(1900, 1, 1)
            dt = base + timedelta(days=int(value) - 2)
            return dt.strftime("%Y%m%d")
        except (OverflowError, ValueError):
            return None

    # 文字列
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None
        # 区切り文字を削除
        digits_only = "".join(c for c in s if c.isdigit())
        if len(digits_only) == 8:  # YYYYMMDD
            return digits_only
        if len(digits_only) == 6:  # YYYYMM → 月初
            return f"{digits_only}01"
        return None

    return None


def _build_header_index(headers: tuple[Any, ...]) -> dict[str, int]:
    """ヘッダ行から列名 → 0-based インデックスのマップを作る。

    同一名の列がある場合（売上明細では「出荷時決済方法」「No 188/190」「None」など）は
    最初に現れた列を採用する。None / 空ヘッダはスキップ。
    """
    idx: dict[str, int] = {}
    for i, h in enumerate(headers):
        if h is None:
            continue
        name = str(h)
        if name in idx:
            # 重複ヘッダは最初の列を優先する（後続列は通常コピーや別用途）
            continue
        idx[name] = i
    return idx


def parse_sales_excel(file_path: str | Path) -> list[dict[str, Any]]:
    """売上明細 Excel を読み込み、必要列のみを辞書のリストに変換する。

    申込月は datetime のまま保持する（calculator 側で正規化）。
    存在しない列はキー自体を省略せず ``None`` を入れる（calculator が defensive lookup する）。
    """
    path = Path(file_path)
    logger.info("売上明細 Excel をパース中: path=%s", path)

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    # シート名は ALL（サンプル） / それ以外でも最初のシートを使う
    ws = wb["ALL"] if "ALL" in wb.sheetnames else wb[wb.sheetnames[0]]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = next(rows_iter)
    except StopIteration:
        wb.close()
        return []

    hidx = _build_header_index(headers)
    records: list[dict[str, Any]] = []

    for row in rows_iter:
        # 完全に空の行はスキップ
        if all(c is None for c in row):
            continue
        rec: dict[str, Any] = {}
        for col_name in _SALES_COLUMNS_TO_KEEP:
            i = hidx.get(col_name)
            rec[col_name] = row[i] if i is not None and i < len(row) else None
        records.append(rec)

    wb.close()
    logger.info("売上明細パース完了: %d 件", len(records))
    return records


def parse_master_excel(file_path: str | Path) -> dict[str, dict[str, Any]]:
    """取引条件マスタ Excel を読み込み、複合キー → 条件レコードの辞書を返す。

    複合キーはマスタ側で事前計算済みの「キー」列を採用する
    （フォーマット: ``{一次店コード:10}{獲得月YYYYMMDD}{商材}{決済方法}``）。
    キーが空のレコードは生成キーで補完する（一次店コード等から再構築）。
    """
    path = Path(file_path)
    logger.info("取引条件マスタ Excel をパース中: path=%s", path)

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb[wb.sheetnames[0]]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = next(rows_iter)
    except StopIteration:
        wb.close()
        return {}

    hidx = _build_header_index(headers)
    master: dict[str, dict[str, Any]] = {}

    key_col_idx = hidx.get("キー")
    primary_code_idx = hidx.get("一次店コード")
    acq_month_idx = hidx.get("獲得月")
    product_idx = hidx.get("商材")
    payment_idx = hidx.get("決済方法")

    for row in rows_iter:
        if all(c is None for c in row):
            continue

        # 「キー」列を優先採用
        key = row[key_col_idx] if key_col_idx is not None and key_col_idx < len(row) else None

        # 「キー」が空のときは構成要素から再構築
        if not key and None not in (primary_code_idx, acq_month_idx, product_idx, payment_idx):
            primary = row[primary_code_idx]
            month_raw = row[acq_month_idx]
            product = row[product_idx]
            payment = row[payment_idx]
            month_str = normalize_yyyymmdd(month_raw)
            if all(v is not None for v in (primary, month_str, product, payment)):
                key = f"{primary}{month_str}{product}{payment}"

        if not key:
            # 突合不能なレコードはスキップ
            continue

        key_str = str(key)

        rec: dict[str, Any] = {}
        for col_name in _MASTER_COLUMNS_TO_KEEP:
            i = hidx.get(col_name)
            rec[col_name] = row[i] if i is not None and i < len(row) else None
        master[key_str] = rec

    wb.close()
    logger.info("マスタパース完了: %d 件のキーを構築", len(master))
    return master
